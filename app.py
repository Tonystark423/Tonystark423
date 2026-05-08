"""Stark Financial Holdings LLC — Asset Ledger."""

import csv
import io
import os
import sqlite3
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import wraps

from dotenv import load_dotenv
from flask import Flask, Response, g, jsonify, render_template, request

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "changeme")

DB_PATH = os.getenv("DB_PATH", "ledger.db")
LEDGER_USER = os.getenv("LEDGER_USER", "admin")
LEDGER_PASS = os.getenv("LEDGER_PASS", "changeme")

COLUMNS = [
    "id", "asset_name", "category", "subcategory", "description",
    "quantity", "unit", "estimated_value", "acquisition_date",
    "custodian", "beneficial_owner", "status", "notes",
    "created_at", "updated_at",
]

WRITABLE_COLUMNS = [c for c in COLUMNS if c not in ("id", "created_at", "updated_at")]

ASSET_NAME_MAX_LENGTH = 100

# Batch signing log columns
SIGNING_COLUMNS = [
    "id", "batch_ref", "asset_category", "signer", "num_items",
    "total_value", "currency", "status", "notes", "signed_at",
    "created_at", "updated_at",
]
SIGNING_WRITABLE = [c for c in SIGNING_COLUMNS if c not in ("id", "created_at", "updated_at")]


def validate_fields(fields: dict) -> tuple[dict, str | None]:
    """
    Validate and sanitize writable asset fields.
    Returns (sanitized_fields, error_message_or_None).

    Rules:
      1. Precision gate: estimated_value must be > 0 when provided
      2. Quantity integrity: quantity must be > 0 when provided
      3. Name sanitization: strip whitespace, cap at ASSET_NAME_MAX_LENGTH
    """
    if "estimated_value" in fields and fields["estimated_value"] is not None:
        try:
            # Use Decimal(str(...)) not Decimal(float) — Decimal(0.29) inherits
            # the float representation error; Decimal("0.29") is exact.
            val = Decimal(str(fields["estimated_value"]))
        except InvalidOperation:
            return fields, "estimated_value must be a number"
        if val <= 0:
            return fields, "estimated_value must be greater than zero"
        # Quantize to 4 decimal places before storing — prevents drift at
        # the boundary. Store as str so SQLite TEXT affinity preserves exact
        # representation; the JSON response converts back via float().
        fields["estimated_value"] = str(val.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))

    if "quantity" in fields and fields["quantity"] is not None:
        try:
            qty = Decimal(str(fields["quantity"]))
        except InvalidOperation:
            return fields, "quantity must be a number"
        if qty <= 0:
            return fields, "quantity must be greater than zero"
        fields["quantity"] = str(qty.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))

    if "asset_name" in fields and fields["asset_name"] is not None:
        fields["asset_name"] = str(fields["asset_name"]).strip()[:ASSET_NAME_MAX_LENGTH]
        if not fields["asset_name"]:
            return fields, "asset_name cannot be blank"

    return fields, None


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or auth.username != LEDGER_USER or auth.password != LEDGER_PASS:
            return Response(
                "Authentication required.",
                401,
                {"WWW-Authenticate": 'Basic realm="Ledger"'},
            )
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
@require_auth
def index():
    return render_template("index.html")


@app.route("/api/assets", methods=["GET"])
@require_auth
def list_assets():
    db = get_db()
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    status = request.args.get("status", "").strip()
    limit = min(int(request.args.get("limit", 200)), 1000)
    offset = int(request.args.get("offset", 0))

    params = []

    if q:
        sql = """
            SELECT a.*
            FROM assets a
            JOIN assets_fts f ON a.id = f.rowid
            WHERE assets_fts MATCH ?
        """
        params.append(q + "*")
        if category:
            sql += " AND a.category = ?"
            params.append(category)
        if status:
            sql += " AND a.status = ?"
            params.append(status)
        sql += " ORDER BY rank LIMIT ? OFFSET ?"
    else:
        sql = "SELECT * FROM assets WHERE 1=1"
        if category:
            sql += " AND category = ?"
            params.append(category)
        if status:
            sql += " AND status = ?"
            params.append(status)
        sql += " ORDER BY created_at DESC LIMIT ? OFFSET ?"

    params += [limit, offset]
    rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/assets", methods=["POST"])
@require_auth
def create_asset():
    data = request.get_json(force=True)
    fields = {k: data[k] for k in WRITABLE_COLUMNS if k in data}
    if "asset_name" not in fields or "category" not in fields:
        return jsonify({"error": "asset_name and category are required"}), 400
    fields, err = validate_fields(fields)
    if err:
        return jsonify({"error": err}), 400

    cols = ", ".join(fields.keys())
    placeholders = ", ".join("?" for _ in fields)
    db = get_db()
    try:
        cur = db.execute(
            f"INSERT INTO assets ({cols}) VALUES ({placeholders})",
            list(fields.values()),
        )
    except sqlite3.IntegrityError as e:
        return jsonify({"error": str(e)}), 400
    db.commit()
    row = db.execute("SELECT * FROM assets WHERE id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/assets/<int:asset_id>", methods=["GET"])
@require_auth
def get_asset(asset_id):
    row = get_db().execute("SELECT * FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if row is None:
        return jsonify({"error": "Not found"}), 404
    return jsonify(dict(row))


@app.route("/api/assets/<int:asset_id>", methods=["PUT"])
@require_auth
def update_asset(asset_id):
    db = get_db()
    if db.execute("SELECT id FROM assets WHERE id = ?", (asset_id,)).fetchone() is None:
        return jsonify({"error": "Not found"}), 404

    data = request.get_json(force=True)
    fields = {k: data[k] for k in WRITABLE_COLUMNS if k in data}
    if not fields:
        return jsonify({"error": "No valid fields provided"}), 400
    fields, err = validate_fields(fields)
    if err:
        return jsonify({"error": err}), 400

    set_clause = ", ".join(f"{k} = ?" for k in fields)
    set_clause += ", updated_at = datetime('now')"
    db.execute(
        f"UPDATE assets SET {set_clause} WHERE id = ?",
        list(fields.values()) + [asset_id],
    )
    db.commit()
    row = db.execute("SELECT * FROM assets WHERE id = ?", (asset_id,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/assets/<int:asset_id>", methods=["DELETE"])
@require_auth
def delete_asset(asset_id):
    db = get_db()
    if db.execute("SELECT id FROM assets WHERE id = ?", (asset_id,)).fetchone() is None:
        return jsonify({"error": "Not found"}), 404
    db.execute("DELETE FROM assets WHERE id = ?", (asset_id,))
    db.commit()
    return "", 204


@app.route("/api/export", methods=["GET"])
@require_auth
def export_csv():
    db = get_db()
    rows = db.execute("SELECT * FROM assets ORDER BY category, asset_name").fetchall()

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=COLUMNS)
    writer.writeheader()
    for row in rows:
        writer.writerow(dict(row))

    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=ledger_export.csv"},
    )


@app.route("/api/import/csv", methods=["POST"])
@require_auth
def import_csv():
    """Upload a CSV file and upsert rows into the ledger.

    Multipart form field: file (the CSV)
    Optional query params:
      beneficial_owner  stamped on rows with no owner column
      custodian         stamped on rows with no custodian column

    Returns:
      { "processed": N, "inserted": N, "skipped": N, "errors": [...] }
    """
    if "file" not in request.files:
        return jsonify({"error": "No file field in request"}), 400

    upload = request.files["file"]
    if not upload.filename or not upload.filename.lower().endswith(".csv"):
        return jsonify({"error": "File must be a .csv"}), 400

    try:
        from ledger_processor import process_csv  # lazy import — optional dependency
    except ImportError:
        return jsonify({"error": "pandas not installed. Run: pip install pandas"}), 503

    owner    = request.args.get("beneficial_owner", os.getenv("LEDGER_BENEFICIAL_OWNER", ""))
    custodian = request.args.get("custodian", "")

    stream = upload.stream
    result = process_csv(stream, get_db(), default_beneficial_owner=owner,
                         default_custodian=custodian)

    status_code = 200 if not result["errors"] else 207  # 207 Multi-Status = partial success
    return jsonify(result), status_code


@app.route("/api/reports/expenses", methods=["GET"])
@require_auth
def report_expenses():
    """Return a PNG bar chart of expenses by spend category.

    Query params:
      after   ISO date — only rows with acquisition_date >= after
      before  ISO date — only rows with acquisition_date <= before

    The chart groups assets where subcategory contains 'debit' or 'expense'
    and applies keyword categorization to the description field.
    """
    try:
        from ledger_processor import categorize_transaction, create_visuals
    except ImportError:
        return jsonify({"error": "pandas/matplotlib not installed"}), 503

    db = get_db()
    sql = "SELECT description, estimated_value, subcategory, acquisition_date FROM assets WHERE 1=1"
    params: list = []

    after  = request.args.get("after", "")
    before = request.args.get("before", "")
    if after:
        sql += " AND acquisition_date >= ?"
        params.append(after)
    if before:
        sql += " AND acquisition_date <= ?"
        params.append(before)

    rows = db.execute(sql, params).fetchall()
    if not rows:
        return jsonify({"error": "No data in ledger"}), 404

    import tempfile
    import pandas as _pd

    df = _pd.DataFrame([dict(r) for r in rows])
    df["Amount"]   = _pd.to_numeric(df["estimated_value"], errors="coerce").fillna(0.0)
    df["Type"]     = df["subcategory"].fillna("").apply(
        lambda s: "Expense" if any(k in s.lower() for k in ("debit", "expense")) else "Income"
    )
    df["Category"] = df["description"].fillna("").apply(categorize_transaction)

    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp_path = tmp.name

    from ledger_processor import BUDGET_LIMITS
    try:
        create_visuals(df, output_path=tmp_path,
                       type_col="Type", category_col="Category", amount_col="Amount",
                       budget_limits=BUDGET_LIMITS)
        with open(tmp_path, "rb") as fh:
            png_bytes = fh.read()
    finally:
        os.unlink(tmp_path)

    return Response(png_bytes, mimetype="image/png",
                    headers={"Content-Disposition": "inline; filename=expense_report.png"})


@app.route("/api/export/excel", methods=["GET"])
@require_auth
def export_excel():
    """Download a comprehensive 8-sheet Excel workbook.

    Ledger sheets (4):
      Raw Data       — every asset row
      Monthly Pivot  — Income / Expense / Net by calendar month
      Budget Check   — actual spend vs. limits per category
      Category Rollup— total spend per keyword category

    Tax sheets (4, OBBBA-aware):
      Summary        — net liability, total proceeds, deductions, savings
      Capital Gains  — per-asset gains with short/long-term classification
      Deductions     — Section 179 + 100% bonus depreciation line items
      Tax Hacks      — all 6 OBBBA optimisation strategies

    Query params:
      after   ISO date — restrict ledger rows included
      before  ISO date — restrict ledger rows included
      year    int      — tax year for the tax sheets (default: current year)
    """
    try:
        from ledger_processor import categorize_transaction, export_excel as _export_ledger, BUDGET_LIMITS
        import pandas as _pd
        import openpyxl
        from copy import copy as _copy
    except ImportError:
        return jsonify({"error": "pandas/openpyxl not installed"}), 503

    try:
        from tax_engine import generate_tax_report, export_tax_excel
    except ImportError:
        return jsonify({"error": "tax_engine module not available"}), 503

    import datetime as _dt
    import tempfile

    db   = get_db()
    year = int(request.args.get("year", _dt.date.today().year))

    # ── Ledger data ──────────────────────────────────────────────────────
    sql    = "SELECT * FROM assets WHERE 1=1"
    params: list = []
    after  = request.args.get("after", "")
    before = request.args.get("before", "")
    if after:
        sql += " AND acquisition_date >= ?"
        params.append(after)
    if before:
        sql += " AND acquisition_date <= ?"
        params.append(before)
    rows = db.execute(sql, params).fetchall()

    # ── Tax report (always generated) ────────────────────────────────────
    report = generate_tax_report(db, tax_year=year)
    report["filing_metadata"] = {
        "entity_name": "Stark Financial Holdings LLC",
        "filed_at":    _dt.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "status":      "draft",
    }
    tax_xlsx = export_tax_excel(report)

    if not rows:
        # No ledger rows — return tax-only workbook rather than 404
        filename = f"stark_financial_report_{year}.xlsx"
        return Response(
            tax_xlsx,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    df = _pd.DataFrame([dict(r) for r in rows])
    df["Amount"]   = _pd.to_numeric(df["estimated_value"], errors="coerce").fillna(0.0)
    df["Type"]     = df["subcategory"].fillna("").apply(
        lambda s: "Expense" if any(k in s.lower() for k in ("debit", "expense")) else "Income"
    )
    df["Category"] = df["description"].fillna("").apply(categorize_transaction)

    # ── Build combined workbook ───────────────────────────────────────────
    ledger_fd, ledger_path = tempfile.mkstemp(suffix=".xlsx")
    tax_fd,    tax_path    = tempfile.mkstemp(suffix=".xlsx")
    try:
        _export_ledger(df, output_path=ledger_path, budget_limits=BUDGET_LIMITS)
        with open(tax_path, "wb") as fh:
            fh.write(tax_xlsx)

        wb_ledger = openpyxl.load_workbook(ledger_path)
        wb_tax    = openpyxl.load_workbook(tax_path)

        combined = openpyxl.Workbook()
        combined.remove(combined.active)   # drop default blank sheet

        for wb in (wb_ledger, wb_tax):
            for name in wb.sheetnames:
                src = wb[name]
                dst = combined.create_sheet(name)
                for row in src.iter_rows():
                    for cell in row:
                        new_cell = dst.cell(
                            row=cell.row, column=cell.column, value=cell.value
                        )
                        if cell.has_style:
                            new_cell.font      = _copy(cell.font)
                            new_cell.fill      = _copy(cell.fill)
                            new_cell.alignment = _copy(cell.alignment)
                for col, dim in src.column_dimensions.items():
                    dst.column_dimensions[col].width = dim.width

        buf = io.BytesIO()
        combined.save(buf)
        xlsx_bytes = buf.getvalue()

    finally:
        os.close(ledger_fd); os.unlink(ledger_path)
        os.close(tax_fd);    os.unlink(tax_path)

    filename = f"stark_financial_report_{year}.xlsx"
    return Response(
        xlsx_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/api/reports/budget", methods=["GET"])
@require_auth
def report_budget():
    """Return a JSON budget check — actual spend vs. limits per category.

    Query params:
      after   ISO date — restrict to rows with acquisition_date >= after
      before  ISO date — restrict to rows with acquisition_date <= before

    Response (200):
      {
        "summary": {"over": 1, "warning": 0, "ok": 6},
        "items": [
          {"category": "Technology", "limit": 5000.0, "actual": 6200.0,
           "variance": -1200.0, "status": "over"},
          ...
        ]
      }
    """
    try:
        from ledger_processor import categorize_transaction, check_budgets
    except ImportError:
        return jsonify({"error": "pandas not installed"}), 503

    import pandas as _pd

    db = get_db()
    sql = "SELECT description, estimated_value, subcategory FROM assets WHERE 1=1"
    params: list = []

    after  = request.args.get("after", "")
    before = request.args.get("before", "")
    if after:
        sql += " AND acquisition_date >= ?"
        params.append(after)
    if before:
        sql += " AND acquisition_date <= ?"
        params.append(before)

    rows = db.execute(sql, params).fetchall()
    if not rows:
        return jsonify({"summary": {"over": 0, "warning": 0, "ok": 0}, "items": []})

    df = _pd.DataFrame([dict(r) for r in rows])
    df["Amount"]   = _pd.to_numeric(df["estimated_value"], errors="coerce").fillna(0.0)
    df["Type"]     = df["subcategory"].fillna("").apply(
        lambda s: "Expense" if any(k in s.lower() for k in ("debit", "expense")) else "Income"
    )
    df["Category"] = df["description"].fillna("").apply(categorize_transaction)

    items = check_budgets(df, type_col="Type", category_col="Category", amount_col="Amount")
    summary = {
        "over":    sum(1 for i in items if i["status"] == "over"),
        "warning": sum(1 for i in items if i["status"] == "warning"),
        "ok":      sum(1 for i in items if i["status"] == "ok"),
    }
    return jsonify({"summary": summary, "items": items})


# ---------------------------------------------------------------------------
# Tax Filing — "Tax Hack"
# ---------------------------------------------------------------------------

@app.route("/api/tax/summary", methods=["GET"])
@require_auth
def tax_summary():
    """Return a JSON tax summary computed from existing assets.

    Query params:
      year  int — tax year label (default: current calendar year)

    Response 200:
      {
        "tax_year": 2025,
        "generated_at": "...",
        "disclaimer": "...",
        "capital_gains": [...],
        "deductions": [...],
        "hacks": [...],
        "summary": { ... }
      }
    """
    try:
        from tax_engine import generate_tax_report
    except ImportError:
        return jsonify({"error": "tax_engine module not available"}), 503

    import datetime as _dt
    year = int(request.args.get("year", _dt.date.today().year))
    report = generate_tax_report(get_db(), tax_year=year)
    return jsonify(report)


@app.route("/api/tax/file", methods=["POST"])
@require_auth
def tax_file():
    """Generate and download a tax filing workbook (Excel).

    Optional JSON body:
      { "year": 2025, "entity_name": "...", "preparer": "..." }

    Response 200: .xlsx attachment with four sheets:
      Summary, Capital Gains, Deductions, Tax Hacks
    """
    try:
        from tax_engine import generate_tax_report, export_tax_excel
    except ImportError:
        return jsonify({"error": "tax_engine module not available"}), 503

    import datetime as _dt

    data = request.get_json(silent=True) or {}
    year        = int(data.get("year", _dt.date.today().year))
    entity_name = data.get("entity_name", "Stark Financial Holdings LLC")
    preparer    = data.get("preparer", "")

    report = generate_tax_report(get_db(), tax_year=year)
    report["filing_metadata"] = {
        "entity_name": entity_name,
        "preparer":    preparer,
        "filed_at":    _dt.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "status":      "draft",
    }

    xlsx_bytes = export_tax_excel(report)
    filename   = f"stark_tax_filing_{year}.xlsx"
    return Response(
        xlsx_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/api/sync/starkbank", methods=["POST"])
@require_auth
def sync_starkbank():
    """Pull live transactions from Stark Bank and upsert into the ledger.

    Optional JSON body:
      { "limit": 50, "after": "2024-01-01", "before": "2024-12-31" }

    Returns:
      { "inserted": N, "skipped": N }
    """
    missing = [v for v in ("STARKBANK_ENVIRONMENT", "STARKBANK_PROJECT_ID", "STARKBANK_PRIVATE_KEY")
               if not os.getenv(v)]
    if missing:
        return jsonify({"error": f"Missing env vars: {', '.join(missing)}"}), 503

    try:
        from starkbank_sync import sync_transactions  # lazy import — optional dependency
    except ImportError:
        return jsonify({"error": "starkbank package not installed. Run: pip install starkbank"}), 503

    data    = request.get_json(silent=True) or {}
    limit   = min(int(data.get("limit", 100)), 1000)
    after   = data.get("after")
    before  = data.get("before")

    try:
        inserted, skipped = sync_transactions(get_db(), limit=limit, after=after, before=before)
    except Exception as exc:  # noqa: BLE001
        return jsonify({"error": str(exc)}), 502

    return jsonify({"inserted": inserted, "skipped": skipped})


# ---------------------------------------------------------------------------
# Bankruptcy claims
# ---------------------------------------------------------------------------

CLAIM_COLUMNS = [
    "id", "claimant_name", "case_number", "court", "trustee_name",
    "trustee_contact", "source", "claim_type", "claimed_value",
    "recovered_value", "currency", "status", "recovered_asset_id",
    "filing_date", "recovery_date", "notes", "created_at", "updated_at",
]
CLAIM_WRITABLE = [c for c in CLAIM_COLUMNS if c not in ("id", "created_at", "updated_at")]


def _validate_claim(fields: dict) -> tuple[dict, str | None]:
    for money_field in ("claimed_value", "recovered_value"):
        if money_field in fields and fields[money_field] is not None:
            try:
                val = Decimal(str(fields[money_field]))
            except InvalidOperation:
                return fields, f"{money_field} must be a number"
            if val < 0:
                return fields, f"{money_field} cannot be negative"
            fields[money_field] = str(val.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))

    if "claimant_name" in fields and fields["claimant_name"] is not None:
        fields["claimant_name"] = str(fields["claimant_name"]).strip()
        if not fields["claimant_name"]:
            return fields, "claimant_name cannot be blank"

    return fields, None


@app.route("/api/claims", methods=["GET"])
@require_auth
def list_claims():
    """List bankruptcy claims.

    Query params:
      claimant   filter by claimant_name (partial, case-insensitive)
      status     identified / filed / pending_recovery / recovered / closed
      source     PACER / State Unclaimed Property / ...
      limit      max rows (default 200, max 1000)
      offset     pagination offset
    """
    db = get_db()
    sql = "SELECT * FROM bankruptcy_claims WHERE 1=1"
    params: list = []

    claimant = request.args.get("claimant", "").strip()
    status   = request.args.get("status",   "").strip()
    source   = request.args.get("source",   "").strip()
    limit    = min(int(request.args.get("limit",  200)), 1000)
    offset   = int(request.args.get("offset", 0))

    if claimant:
        sql += " AND claimant_name LIKE ?"
        params.append(f"%{claimant}%")
    if status:
        sql += " AND status = ?"
        params.append(status)
    if source:
        sql += " AND source = ?"
        params.append(source)

    sql += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
    params += [limit, offset]
    rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/claims", methods=["POST"])
@require_auth
def create_claim():
    """Record a new bankruptcy claim.

    Required JSON body: claimant_name
    Optional: case_number, court, trustee_name, trustee_contact, source,
              claim_type, claimed_value, currency, status, filing_date, notes
    """
    data = request.get_json(force=True)
    fields = {k: data[k] for k in CLAIM_WRITABLE if k in data}
    if "claimant_name" not in fields:
        return jsonify({"error": "claimant_name is required"}), 400
    fields, err = _validate_claim(fields)
    if err:
        return jsonify({"error": err}), 400

    cols = ", ".join(fields.keys())
    placeholders = ", ".join("?" for _ in fields)
    db = get_db()
    try:
        cur = db.execute(
            f"INSERT INTO bankruptcy_claims ({cols}) VALUES ({placeholders})",
            list(fields.values()),
        )
    except sqlite3.IntegrityError as e:
        return jsonify({"error": str(e)}), 400
    db.commit()
    row = db.execute("SELECT * FROM bankruptcy_claims WHERE id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/claims/<int:claim_id>", methods=["GET"])
@require_auth
def get_claim(claim_id):
    row = get_db().execute(
        "SELECT * FROM bankruptcy_claims WHERE id = ?", (claim_id,)
    ).fetchone()
    if row is None:
        return jsonify({"error": "Not found"}), 404
    return jsonify(dict(row))


@app.route("/api/claims/<int:claim_id>", methods=["PUT"])
@require_auth
def update_claim(claim_id):
    """Update a claim — advance status, record recovery details, link asset.

    To mark a claim recovered:
      { "status": "recovered", "recovered_value": "5000.00",
        "recovery_date": "2025-04-09", "recovered_asset_id": 42 }
    """
    db = get_db()
    if db.execute("SELECT id FROM bankruptcy_claims WHERE id = ?", (claim_id,)).fetchone() is None:
        return jsonify({"error": "Not found"}), 404

    data = request.get_json(force=True)
    fields = {k: data[k] for k in CLAIM_WRITABLE if k in data}
    if not fields:
        return jsonify({"error": "No valid fields provided"}), 400
    fields, err = _validate_claim(fields)
    if err:
        return jsonify({"error": err}), 400

    set_clause = ", ".join(f"{k} = ?" for k in fields)
    set_clause += ", updated_at = datetime('now')"
    db.execute(
        f"UPDATE bankruptcy_claims SET {set_clause} WHERE id = ?",
        list(fields.values()) + [claim_id],
    )
    db.commit()
    row = db.execute("SELECT * FROM bankruptcy_claims WHERE id = ?", (claim_id,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/claims/summary", methods=["GET"])
@require_auth
def claims_summary():
    """Return bankruptcy_summary view — totals grouped by claimant, status, source."""
    rows = get_db().execute("SELECT * FROM bankruptcy_summary ORDER BY claimant_name, status").fetchall()
    return jsonify([dict(r) for r in rows])


# ---------------------------------------------------------------------------
# Portfolio summary — crypto + equities + all holdings grouped by category
# ---------------------------------------------------------------------------

@app.route("/api/portfolio/summary", methods=["GET"])
@require_auth
def portfolio_summary():
    """Return holdings grouped by category, subcategory, status, and owner.

    Optional query params:
      owner     filter by beneficial_owner (e.g. 'All-Star Financial Holdings')
      category  filter by asset category
      status    filter by asset status (active / sold / pending)

    Response (200):
      [
        { "category": "Cryptocurrency", "subcategory": "BTC", "status": "active",
          "beneficial_owner": "All-Star Financial Holdings",
          "asset_count": 3, "total_value_usd": 125000.0, "unit": "BTC" },
        ...
      ]
    """
    db = get_db()
    sql = "SELECT * FROM portfolio_summary WHERE 1=1"
    params: list = []

    owner = request.args.get("owner", "").strip()
    category = request.args.get("category", "").strip()
    status = request.args.get("status", "").strip()

    if owner:
        sql += " AND beneficial_owner = ?"
        params.append(owner)
    if category:
        sql += " AND category = ?"
        params.append(category)
    if status:
        sql += " AND status = ?"
        params.append(status)

    sql += " ORDER BY category, subcategory"
    rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


# ---------------------------------------------------------------------------
# Batch signing log
# ---------------------------------------------------------------------------

def _validate_signing(fields: dict) -> tuple[dict, str | None]:
    """Validate writable batch_signings fields."""
    if "total_value" in fields and fields["total_value"] is not None:
        try:
            val = Decimal(str(fields["total_value"]))
        except InvalidOperation:
            return fields, "total_value must be a number"
        if val <= 0:
            return fields, "total_value must be greater than zero"
        fields["total_value"] = str(val.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))

    if "num_items" in fields and fields["num_items"] is not None:
        try:
            n = int(fields["num_items"])
        except (TypeError, ValueError):
            return fields, "num_items must be an integer"
        if n < 0:
            return fields, "num_items cannot be negative"
        fields["num_items"] = n

    if "batch_ref" in fields and fields["batch_ref"] is not None:
        fields["batch_ref"] = str(fields["batch_ref"]).strip()
        if not fields["batch_ref"]:
            return fields, "batch_ref cannot be blank"

    return fields, None


@app.route("/api/signings", methods=["GET"])
@require_auth
def list_signings():
    """List batch signing records.

    Query params:
      status         pending / signed / failed / cancelled
      asset_category Cryptocurrency / Securities & Commodities / ...
      limit          max rows (default 200, max 1000)
      offset         pagination offset
    """
    db = get_db()
    sql = "SELECT * FROM batch_signings WHERE 1=1"
    params: list = []

    status = request.args.get("status", "").strip()
    category = request.args.get("asset_category", "").strip()
    limit = min(int(request.args.get("limit", 200)), 1000)
    offset = int(request.args.get("offset", 0))

    if status:
        sql += " AND status = ?"
        params.append(status)
    if category:
        sql += " AND asset_category = ?"
        params.append(category)

    sql += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
    params += [limit, offset]
    rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/signings", methods=["POST"])
@require_auth
def create_signing():
    """Record a new batch signing event.

    Required JSON body fields: batch_ref
    Optional: asset_category, signer, num_items, total_value, currency,
              status, notes, signed_at
    """
    data = request.get_json(force=True)
    fields = {k: data[k] for k in SIGNING_WRITABLE if k in data}
    if "batch_ref" not in fields:
        return jsonify({"error": "batch_ref is required"}), 400
    fields, err = _validate_signing(fields)
    if err:
        return jsonify({"error": err}), 400

    cols = ", ".join(fields.keys())
    placeholders = ", ".join("?" for _ in fields)
    db = get_db()
    try:
        cur = db.execute(
            f"INSERT INTO batch_signings ({cols}) VALUES ({placeholders})",
            list(fields.values()),
        )
    except sqlite3.IntegrityError as e:
        return jsonify({"error": str(e)}), 400
    db.commit()
    row = db.execute("SELECT * FROM batch_signings WHERE id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/signings/<int:signing_id>", methods=["GET"])
@require_auth
def get_signing(signing_id):
    row = get_db().execute(
        "SELECT * FROM batch_signings WHERE id = ?", (signing_id,)
    ).fetchone()
    if row is None:
        return jsonify({"error": "Not found"}), 404
    return jsonify(dict(row))


@app.route("/api/signings/<int:signing_id>", methods=["PUT"])
@require_auth
def update_signing(signing_id):
    """Update a batch signing record (e.g. advance status from pending -> signed)."""
    db = get_db()
    if db.execute("SELECT id FROM batch_signings WHERE id = ?", (signing_id,)).fetchone() is None:
        return jsonify({"error": "Not found"}), 404

    data = request.get_json(force=True)
    fields = {k: data[k] for k in SIGNING_WRITABLE if k in data}
    if not fields:
        return jsonify({"error": "No valid fields provided"}), 400
    fields, err = _validate_signing(fields)
    if err:
        return jsonify({"error": err}), 400

    set_clause = ", ".join(f"{k} = ?" for k in fields)
    set_clause += ", updated_at = datetime('now')"
    db.execute(
        f"UPDATE batch_signings SET {set_clause} WHERE id = ?",
        list(fields.values()) + [signing_id],
    )
    db.commit()
    row = db.execute("SELECT * FROM batch_signings WHERE id = ?", (signing_id,)).fetchone()
    return jsonify(dict(row))


if __name__ == "__main__":
    app.run(debug=False)
