"""Tax filing engine for Stark Financial Holdings LLC.

Analyses existing asset data to:
  - Calculate realised capital gains / losses (sold assets)
  - Identify deductions (Section 179, business expenses)
  - Apply tax optimisation hacks (loss harvesting, deduction maximisation)
  - Generate a structured tax report and exportable Excel workbook
"""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Federal long-term capital gains rate for high-income filers
_LT_RATE = Decimal("0.20")
# Federal short-term capital gains rate (ordinary income, top bracket)
_ST_RATE = Decimal("0.37")
# Net investment income tax (NIIT) rate
_NIIT_RATE = Decimal("0.038")

# IRS: "more than 1 year" = strictly greater than 365 days for long-term treatment
_LT_HOLDING_DAYS = 365

# ---------------------------------------------------------------------------
# One Big Beautiful Bill Act (OBBBA / H.R. 1, 2025) provisions
# ---------------------------------------------------------------------------
# Section 179 expensing limit: raised from $1,220,000 → $2,500,000 (base 2025).
# Indexed for inflation — 2026 IRS-adjusted limit: $2,560,000.
_SEC179_LIMIT       = Decimal("2560000.00")
# Section 179 phase-out threshold: raised from $3,050,000 → $4,000,000 (base 2025).
# Indexed for inflation — 2026 IRS-adjusted phase-out: $4,090,000.
_SEC179_PHASE_OUT   = Decimal("4090000.00")
# Bonus depreciation: restored to 100% (was 40% in 2025 under prior law)
# Applies to qualifying property placed in service after 1/19/2025
_BONUS_DEP_RATE     = Decimal("1.00")
_BONUS_DEP_CUTOFF   = "2025-01-19"
# Section 199A pass-through deduction: permanently extended at 20%.
# (OBBBA did NOT raise the rate; it made the 20% deduction permanent.)
_SEC199A_RATE       = Decimal("0.20")
# QOZ incentives extended through 2034
_QOZ_EXTENDED_YEAR  = 2034

# Categories eligible for Section 179 / bonus depreciation (equipment)
_SEC179_CATEGORIES = {"Computer Resources", "Proprietary IP"}

# Heavy vehicle detection — vehicles with GVWR > 6,000 lbs are exempt from
# luxury-auto depreciation caps and qualify for Section 179 / bonus dep.
# Detected via subcategory AND notes keywords written by seed_assets.py.
_HEAVY_VEHICLE_SUBCATS   = {"Ground Transportation", "Executive Transport"}
_HEAVY_VEHICLE_NOTE_KEYS = (
    "GVWR > 6,000",
    "exceeds 6,000",
    "over weight limit",
    "exempt from luxury",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_decimal(value: Any) -> Decimal:
    """Convert a raw value to Decimal, returning 0 on failure."""
    try:
        return Decimal(str(value)).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, TypeError):
        return Decimal("0")


def _holding_days(acquisition_date_str: str | None) -> int | None:
    """Return days held up to today, or None if date is missing/invalid."""
    if not acquisition_date_str:
        return None
    try:
        acq = datetime.strptime(acquisition_date_str[:10], "%Y-%m-%d").date()
        return (date.today() - acq).days
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Core analysis functions
# ---------------------------------------------------------------------------

def get_capital_gains(conn) -> list[dict]:
    """Return a list of realised gain/loss records from sold assets.

    Note: cost basis is not tracked in the ledger schema, so *estimated_value*
    is treated as the proceeds at the time the asset was recorded as sold.
    Each record is flagged ``cost_basis_known: false`` to make this clear.
    """
    rows = conn.execute(
        "SELECT id, asset_name, category, estimated_value, acquisition_date, notes "
        "FROM assets WHERE status = 'sold' ORDER BY acquisition_date"
    ).fetchall()

    gains = []
    for row in rows:
        proceeds = _to_decimal(row["estimated_value"])
        days = _holding_days(row["acquisition_date"])
        if days is not None and days > _LT_HOLDING_DAYS:
            gain_type = "long_term"
            rate = _LT_RATE
        else:
            gain_type = "short_term"
            rate = _ST_RATE

        estimated_tax = (proceeds * rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        gains.append({
            "id": row["id"],
            "asset_name": row["asset_name"],
            "category": row["category"],
            "proceeds": str(proceeds),
            "gain_type": gain_type,
            "holding_days": days,
            "estimated_tax": str(estimated_tax),
            "tax_rate": str(rate),
            "cost_basis_known": False,
            "note": "Cost basis not tracked; proceeds treated as full gain.",
        })

    return gains


def get_deductions(conn) -> list[dict]:
    """Return deduction opportunities from active assets.

    Qualifying assets (processed oldest-first to maximise Section 179 on pre-cutoff
    assets, leaving post-cutoff assets free for unlimited bonus depreciation):

      Equipment  — category in _SEC179_CATEGORIES (Computer Resources, Proprietary IP)
      Heavy veh. — subcategory in _HEAVY_VEHICLE_SUBCATS AND notes signal GVWR > 6,000 lbs
                   (exempt from luxury-auto caps under IRC §179; Tesla Model S excluded)

    Strategy (OBBBA-optimal):
      1. Section 179 — shared $2,500,000 cap across all qualifying asset types.
      2. 100% Bonus Depreciation — unlimited, for assets placed in service after
         _BONUS_DEP_CUTOFF (2025-01-19). No dollar cap.
    """
    note_filter = " OR ".join(
        f"COALESCE(notes,'') LIKE ?" for _ in _HEAVY_VEHICLE_NOTE_KEYS
    )
    subcat_filter = ",".join("?" for _ in _HEAVY_VEHICLE_SUBCATS)

    rows = conn.execute(
        f"""
        SELECT id, asset_name, category, subcategory, estimated_value, acquisition_date, notes
        FROM assets
        WHERE status IN ('active', 'pending')
          AND (
            category IN ({",".join("?" for _ in _SEC179_CATEGORIES)})
            OR (
              subcategory IN ({subcat_filter})
              AND ({note_filter})
            )
          )
        ORDER BY acquisition_date ASC, asset_name
        """,
        [*_SEC179_CATEGORIES, *_HEAVY_VEHICLE_SUBCATS,
         *[f"%{k}%" for k in _HEAVY_VEHICLE_NOTE_KEYS]],
    ).fetchall()

    deductions = []
    sec179_used = Decimal("0")

    for row in rows:
        value = _to_decimal(row["estimated_value"])
        if value <= 0:
            continue

        category  = row["category"] or ""
        subcat    = row["subcategory"] or ""
        acq_date  = (row["acquisition_date"] or "")[:10]
        remaining = value

        is_heavy_vehicle = subcat in _HEAVY_VEHICLE_SUBCATS
        label = "Heavy Vehicle" if is_heavy_vehicle else category

        # ── Section 179 (up to OBBBA cap) ────────────────────────────────
        sec179_available = _SEC179_LIMIT - sec179_used
        if sec179_available > 0:
            deductible_179 = min(remaining, sec179_available)
            sec179_used   += deductible_179
            remaining     -= deductible_179
            deductions.append({
                "id":               row["id"],
                "asset_name":       row["asset_name"],
                "category":         category,
                "deduction_type":   "Section 179",
                "deductible_amount": str(deductible_179.quantize(Decimal("0.01"))),
                "description": (
                    f"{label} qualifies for Section 179 expensing "
                    f"(OBBBA 2025 limit: ${_SEC179_LIMIT:,.2f}/year; "
                    f"phase-out above ${_SEC179_PHASE_OUT:,.2f})."
                    + (" GVWR > 6,000 lbs — exempt from luxury-auto depreciation caps."
                       if is_heavy_vehicle else "")
                ),
            })

        # ── 100% Bonus Depreciation on remaining basis (post-cutoff only) ─
        if remaining > 0 and acq_date >= _BONUS_DEP_CUTOFF:
            deductions.append({
                "id":               row["id"],
                "asset_name":       row["asset_name"],
                "category":         category,
                "deduction_type":   "100% Bonus Depreciation",
                "deductible_amount": str(remaining.quantize(Decimal("0.01"))),
                "description": (
                    f"OBBBA restores 100% first-year bonus depreciation for qualifying "
                    f"property placed in service after {_BONUS_DEP_CUTOFF} — no dollar cap. "
                    f"Full ${remaining:,.2f} remaining basis expensed in year one."
                    + (" GVWR > 6,000 lbs — not subject to luxury-auto limits."
                       if is_heavy_vehicle else "")
                ),
            })

    return deductions


def apply_tax_hacks(gains: list[dict], deductions: list[dict]) -> dict:
    """Apply tax optimisation strategies and return a hacks summary.

    Strategies applied:
      1. Section 179 expensing (OBBBA: $2.5M limit)
      2. Hold-for-long-term rate preference
      3. NIIT reduction via deductions
      4. QOZ deferral (OBBBA: extended through 2034)
      5. 100% Bonus depreciation (OBBBA: restored, no dollar cap)
      6. Section 199A pass-through deduction (OBBBA: 23%)
    """
    hacks = []
    total_savings = Decimal("0")

    total_gains   = sum((Decimal(g["proceeds"]) for g in gains), Decimal("0"))

    # Split deductions by mechanism so each hack carries the right figure
    sec179_total  = sum(
        (Decimal(d["deductible_amount"]) for d in deductions if d.get("deduction_type") == "Section 179"),
        Decimal("0"),
    )
    bonus_dep_total = sum(
        (Decimal(d["deductible_amount"]) for d in deductions if d.get("deduction_type") == "100% Bonus Depreciation"),
        Decimal("0"),
    )
    total_deductions = sec179_total + bonus_dep_total

    # Hack 1 — Section 179 expensing (OBBBA limit: $2,500,000)
    if sec179_total > 0:
        savings_179 = (sec179_total * _ST_RATE).quantize(Decimal("0.01"))
        total_savings += savings_179
        hacks.append({
            "hack": "Section 179 Immediate Expensing (OBBBA: $2.5M limit)",
            "description": (
                f"Deduct ${sec179_total:,.2f} of qualifying assets this tax year "
                f"instead of depreciating over multiple years. "
                f"OBBBA raised the Section 179 limit from $1,220,000 to ${_SEC179_LIMIT:,.2f}."
            ),
            "estimated_savings": str(savings_179),
        })

    # Hack 2 — Long-term holding preference
    short_term = [g for g in gains if g["gain_type"] == "short_term"]
    if short_term:
        st_proceeds = sum(Decimal(g["proceeds"]) for g in short_term)
        rate_diff = _ST_RATE - _LT_RATE
        potential_savings = (st_proceeds * rate_diff).quantize(Decimal("0.01"))
        hacks.append({
            "hack": "Hold for Long-Term Rates",
            "description": (
                f"{len(short_term)} sold asset(s) were held short-term. "
                f"Holding equivalent future positions for 12+ months could save "
                f"~${potential_savings:,.2f} at the {rate_diff*100:.0f}% rate differential."
            ),
            "estimated_savings": str(potential_savings),
        })

    # Hack 3 — NIIT reduction via deductions
    if total_deductions > 0 and total_gains > 0:
        niit_savings = (min(total_deductions, total_gains) * _NIIT_RATE).quantize(Decimal("0.01"))
        total_savings += niit_savings
        hacks.append({
            "hack": "Net Investment Income Tax (NIIT) Reduction",
            "description": (
                f"Business deductions reduce net investment income subject to the 3.8% NIIT. "
                f"Estimated NIIT savings: ${niit_savings:,.2f}."
            ),
            "estimated_savings": str(niit_savings),
        })

    # Hack 4 — QOZ deferral (OBBBA: extended through 2034)
    if total_gains > 0:
        hacks.append({
            "hack": f"Qualified Opportunity Zone (QOZ) Deferral — extended to {_QOZ_EXTENDED_YEAR}",
            "description": (
                "Reinvesting capital gains into a Qualified Opportunity Fund can defer "
                "and potentially reduce taxes on those gains. "
                f"OBBBA extends QOZ incentives through {_QOZ_EXTENDED_YEAR}. "
                "Consult a tax adviser."
            ),
            "estimated_savings": "varies",
        })

    # Hack 5 — 100% Bonus Depreciation (OBBBA: restored, no dollar cap)
    if bonus_dep_total > 0:
        bonus_savings = (bonus_dep_total * _ST_RATE).quantize(Decimal("0.01"))
        total_savings += bonus_savings
        hacks.append({
            "hack": "100% Bonus Depreciation (OBBBA: restored, no dollar cap)",
            "description": (
                f"OBBBA permanently restores 100% first-year bonus depreciation for qualifying "
                f"property placed in service after {_BONUS_DEP_CUTOFF}. "
                f"${bonus_dep_total:,.2f} of post-cutoff Computer Resources / Proprietary IP "
                f"expensed in full — no dollar cap, unlike Section 179's ${_SEC179_LIMIT:,.2f} limit. "
                f"Est. savings: ${bonus_savings:,.2f}."
            ),
            "estimated_savings": str(bonus_savings),
        })
    elif sec179_total > 0:
        hacks.append({
            "hack": "100% Bonus Depreciation (OBBBA: restored, no dollar cap)",
            "description": (
                f"OBBBA permanently restores 100% first-year bonus depreciation for qualifying "
                f"property placed in service after {_BONUS_DEP_CUTOFF}. "
                f"Unlike Section 179 (capped at ${_SEC179_LIMIT:,.2f}), bonus depreciation "
                f"has no dollar limit — consider acquiring new Computer Resources or Proprietary IP "
                f"after {_BONUS_DEP_CUTOFF} to unlock unlimited first-year expensing."
            ),
            "estimated_savings": "varies",
        })

    # Hack 6 — Section 199A pass-through deduction (OBBBA: 20%, now permanent)
    hacks.append({
        "hack": f"Section 199A Pass-Through Deduction (OBBBA: {int(_SEC199A_RATE * 100)}%, permanent)",
        "description": (
            f"OBBBA permanently extends the qualified business income (QBI) deduction at "
            f"{int(_SEC199A_RATE * 100)}% for eligible pass-through entities such as LLCs. "
            f"Without OBBBA this would have expired after 2025. "
            f"Stark Financial Holdings LLC may deduct {int(_SEC199A_RATE * 100)}% of net QBI, "
            f"substantially reducing effective tax rate on pass-through income. "
            f"2026 threshold: $203,000 (single) / $406,000 (MFJ) before W-2 wage limits apply. "
            f"Minimum $400 deduction guaranteed if QBI ≥ $1,000 (OBBBA new provision). "
            f"Consult a tax adviser to confirm eligibility."
        ),
        "estimated_savings": "varies",
    })

    return {
        "hacks": hacks,
        "total_estimated_savings": str(total_savings.quantize(Decimal("0.01"))),
    }


def generate_tax_report(conn, tax_year: int = 2025) -> dict:
    """Compile a full tax report for the given tax year.

    Returns a dict suitable for JSON serialisation containing:
      - tax_year
      - capital_gains  (list from get_capital_gains)
      - deductions     (list from get_deductions)
      - hacks          (dict from apply_tax_hacks)
      - summary        (totals and estimated net liability)
    """
    gains = get_capital_gains(conn)
    deductions = get_deductions(conn)
    hacks_result = apply_tax_hacks(gains, deductions)

    total_proceeds = sum((Decimal(g["proceeds"]) for g in gains), Decimal("0"))
    total_tax_before = sum((Decimal(g["estimated_tax"]) for g in gains), Decimal("0"))
    total_deductible = sum((Decimal(d["deductible_amount"]) for d in deductions), Decimal("0"))

    # Net tax after Section 179 deductions (applied at short-term rate as a proxy)
    deduction_tax_benefit = (total_deductible * _ST_RATE).quantize(Decimal("0.01"))
    net_liability = max(Decimal("0"), total_tax_before - deduction_tax_benefit)

    return {
        "tax_year": tax_year,
        "generated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "disclaimer": (
            "These figures are ESTIMATES based on incomplete data (no cost basis). "
            "Incorporates One Big Beautiful Bill Act (OBBBA / H.R. 1, 2025) provisions: "
            f"Section 179 limit ${_SEC179_LIMIT:,.0f}, 100% bonus depreciation, "
            f"Section 199A at {int(_SEC199A_RATE * 100)}%, QOZ extended to {_QOZ_EXTENDED_YEAR}. "
            "Consult a qualified tax professional before filing."
        ),
        "capital_gains": gains,
        "deductions": deductions,
        "hacks": hacks_result["hacks"],
        "summary": {
            "total_proceeds": str(total_proceeds.quantize(Decimal("0.01"))),
            "estimated_tax_before_deductions": str(total_tax_before),
            "total_deductions": str(total_deductible.quantize(Decimal("0.01"))),
            "deduction_tax_benefit": str(deduction_tax_benefit),
            "estimated_net_liability": str(net_liability),
            "total_estimated_savings": hacks_result["total_estimated_savings"],
        },
    }


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_tax_excel(report: dict) -> bytes:
    """Render the tax report as a multi-sheet Excel workbook and return bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ---- colour palette (dark theme matching the app) ----
    HDR_FILL = PatternFill("solid", fgColor="1A1D27")
    HDR_FONT = Font(bold=True, color="4F8EF7", size=11)
    TITLE_FONT = Font(bold=True, color="E2E4ED", size=13)
    MUTED_FONT = Font(color="8B8FA8", italic=True)

    def _style_header_row(ws, row_idx: int, ncols: int):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")

    def _auto_width(ws):
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 50)

    # ------------------------------------------------------------------ #
    # Sheet 1 — Summary
    # ------------------------------------------------------------------ #
    ws_sum = wb.active
    ws_sum.title = "Summary"
    summary = report["summary"]
    ws_sum["A1"] = f"Stark Financial Holdings LLC — Tax Return {report['tax_year']}"
    ws_sum["A1"].font = TITLE_FONT
    ws_sum["A2"] = report["disclaimer"]
    ws_sum["A2"].font = MUTED_FONT
    ws_sum["A2"].alignment = Alignment(wrap_text=True)
    ws_sum.row_dimensions[2].height = 40

    rows_sum = [
        ("Generated at", report["generated_at"]),
        ("Tax Year", report["tax_year"]),
        ("", ""),
        ("Total Proceeds (Sold Assets)", f"${float(summary['total_proceeds']):,.2f}"),
        ("Est. Tax Before Deductions", f"${float(summary['estimated_tax_before_deductions']):,.2f}"),
        ("Total Deductions", f"${float(summary['total_deductions']):,.2f}"),
        ("Deduction Tax Benefit", f"${float(summary['deduction_tax_benefit']):,.2f}"),
        ("", ""),
        ("ESTIMATED NET LIABILITY", f"${float(summary['estimated_net_liability']):,.2f}"),
        ("Total Estimated Savings", f"${float(summary['total_estimated_savings']):,.2f}"),
    ]
    for i, (label, value) in enumerate(rows_sum, start=4):
        ws_sum.cell(row=i, column=1, value=label)
        ws_sum.cell(row=i, column=2, value=value)
        if label == "ESTIMATED NET LIABILITY":
            ws_sum.cell(row=i, column=1).font = Font(bold=True, color="E05C5C")
            ws_sum.cell(row=i, column=2).font = Font(bold=True, color="E05C5C")
    _auto_width(ws_sum)

    # ------------------------------------------------------------------ #
    # Sheet 2 — Capital Gains
    # ------------------------------------------------------------------ #
    ws_cg = wb.create_sheet("Capital Gains")
    cg_headers = ["Asset Name", "Category", "Proceeds ($)", "Gain Type",
                  "Holding Days", "Est. Tax ($)", "Tax Rate", "Notes"]
    for col, h in enumerate(cg_headers, start=1):
        ws_cg.cell(row=1, column=col, value=h)
    _style_header_row(ws_cg, 1, len(cg_headers))

    for row_i, g in enumerate(report["capital_gains"], start=2):
        ws_cg.cell(row=row_i, column=1, value=g["asset_name"])
        ws_cg.cell(row=row_i, column=2, value=g["category"])
        ws_cg.cell(row=row_i, column=3, value=float(g["proceeds"]))
        ws_cg.cell(row=row_i, column=4, value=g["gain_type"].replace("_", "-"))
        ws_cg.cell(row=row_i, column=5, value=g["holding_days"])
        ws_cg.cell(row=row_i, column=6, value=float(g["estimated_tax"]))
        ws_cg.cell(row=row_i, column=7, value=g["tax_rate"])
        ws_cg.cell(row=row_i, column=8, value=g["note"])
    _auto_width(ws_cg)

    # ------------------------------------------------------------------ #
    # Sheet 3 — Deductions
    # ------------------------------------------------------------------ #
    ws_ded = wb.create_sheet("Deductions")
    ded_headers = ["Asset Name", "Category", "Deduction Type", "Deductible Amount ($)", "Description"]
    for col, h in enumerate(ded_headers, start=1):
        ws_ded.cell(row=1, column=col, value=h)
    _style_header_row(ws_ded, 1, len(ded_headers))

    for row_i, d in enumerate(report["deductions"], start=2):
        ws_ded.cell(row=row_i, column=1, value=d["asset_name"])
        ws_ded.cell(row=row_i, column=2, value=d["category"])
        ws_ded.cell(row=row_i, column=3, value=d["deduction_type"])
        ws_ded.cell(row=row_i, column=4, value=float(d["deductible_amount"]))
        ws_ded.cell(row=row_i, column=5, value=d["description"])
    _auto_width(ws_ded)

    # ------------------------------------------------------------------ #
    # Sheet 4 — Tax Hacks
    # ------------------------------------------------------------------ #
    ws_hacks = wb.create_sheet("Tax Hacks")
    hack_headers = ["Strategy", "Description", "Estimated Savings ($)"]
    for col, h in enumerate(hack_headers, start=1):
        ws_hacks.cell(row=1, column=col, value=h)
    _style_header_row(ws_hacks, 1, len(hack_headers))

    for row_i, h in enumerate(report["hacks"], start=2):
        ws_hacks.cell(row=row_i, column=1, value=h["hack"])
        ws_hacks.cell(row=row_i, column=2, value=h["description"])
        ws_hacks.cell(row=row_i, column=3, value=h["estimated_savings"])
    _auto_width(ws_hacks)

    # ------------------------------------------------------------------ #
    # Sheet 5 — Schedule D (IRS form structure)
    # ------------------------------------------------------------------ #
    ws_sd = wb.create_sheet("Schedule D")
    ws_sd["A1"] = f"Schedule D — Capital Gains and Losses  (Tax Year {report['tax_year']})"
    ws_sd["A1"].font = TITLE_FONT
    ws_sd["A2"] = "Stark Financial Holdings LLC  |  Prepared by Stark Tax Engine"
    ws_sd["A2"].font = MUTED_FONT

    sd_headers = ["(a) Description", "(b) Date Acquired", "(c) Date Sold / Disposed",
                  "(d) Proceeds ($)", "(e) Cost Basis ($)", "(f) Adj ($)", "(g) Gain / Loss ($)",
                  "Gain Type", "Est. Federal Tax ($)"]
    for col, h in enumerate(sd_headers, start=1):
        ws_sd.cell(row=4, column=col, value=h)
    _style_header_row(ws_sd, 4, len(sd_headers))

    short_gains = [g for g in report["capital_gains"] if g["gain_type"] == "short_term"]
    long_gains  = [g for g in report["capital_gains"] if g["gain_type"] == "long_term"]

    def _write_sd_section(ws, title, gains_list, start_row):
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, color="4F8EF7")
        r = start_row + 1
        for g in gains_list:
            proceeds = float(g["proceeds"])
            ws.cell(row=r, column=1, value=g["asset_name"])
            ws.cell(row=r, column=2, value=g.get("acquisition_date", "N/A"))
            ws.cell(row=r, column=3, value="See ledger")
            ws.cell(row=r, column=4, value=proceeds)
            ws.cell(row=r, column=5, value="N/A — basis not tracked")
            ws.cell(row=r, column=6, value=0)
            ws.cell(row=r, column=7, value=proceeds)   # treated as full gain
            ws.cell(row=r, column=8, value=g["gain_type"].replace("_", "-"))
            ws.cell(row=r, column=9, value=float(g["estimated_tax"]))
            r += 1
        subtotal = sum(float(g["proceeds"]) for g in gains_list)
        ws.cell(row=r, column=1, value=f"Subtotal ({title})").font = Font(bold=True)
        ws.cell(row=r, column=4, value=subtotal).font = Font(bold=True)
        ws.cell(row=r, column=9,
                value=sum(float(g["estimated_tax"]) for g in gains_list)).font = Font(bold=True)
        return r + 2

    row_ptr = 5
    row_ptr = _write_sd_section(ws_sd, "Part I — Short-Term (held ≤ 1 year; rate: 37%)", short_gains, row_ptr)
    row_ptr = _write_sd_section(ws_sd, "Part II — Long-Term (held > 1 year; rate: 20%)", long_gains, row_ptr)

    total_tax = sum(float(g["estimated_tax"]) for g in report["capital_gains"])
    ws_sd.cell(row=row_ptr, column=1, value="NET CAPITAL GAINS TAX (Federal, est.)").font = Font(bold=True, color="E05C5C")
    ws_sd.cell(row=row_ptr, column=9, value=total_tax).font = Font(bold=True, color="E05C5C")
    ws_sd.cell(row=row_ptr + 1, column=1,
               value="* Cost basis not tracked in ledger. Proceeds treated as full gain. Update with actual basis before filing."
               ).font = MUTED_FONT
    _auto_width(ws_sd)

    # ------------------------------------------------------------------ #
    # Sheet 6 — Form 4562 (Depreciation & Amortisation)
    # ------------------------------------------------------------------ #
    ws_4562 = wb.create_sheet("Form 4562")
    ws_4562["A1"] = f"Form 4562 — Depreciation and Amortization  (Tax Year {report['tax_year']})"
    ws_4562["A1"].font = TITLE_FONT
    ws_4562["A2"] = "Stark Financial Holdings LLC  |  EIN: XX-XXXXXXX  |  OBBBA provisions applied"
    ws_4562["A2"].font = MUTED_FONT

    sec179_items  = [d for d in report["deductions"] if d["deduction_type"] == "Section 179"]
    bonus_items   = [d for d in report["deductions"] if d["deduction_type"] == "100% Bonus Depreciation"]
    sec179_total  = sum(float(d["deductible_amount"]) for d in sec179_items)
    bonus_total   = sum(float(d["deductible_amount"]) for d in bonus_items)

    ws_4562["A4"] = "PART I — Election to Expense Certain Property (Section 179)"
    ws_4562["A4"].font = Font(bold=True, color="4F8EF7")
    part1 = [
        ("Line 1",  "Maximum Section 179 amount (OBBBA 2026 inflation-adjusted)", f"${_SEC179_LIMIT:,.2f}"),
        ("Line 2",  "Phase-out threshold (OBBBA 2026 inflation-adjusted)",         f"${_SEC179_PHASE_OUT:,.2f}"),
        ("Line 6",  "Total cost of Section 179 property placed in service",        f"${sec179_total:,.2f}"),
        ("Line 12", "Section 179 expense deduction this year",                     f"${sec179_total:,.2f}"),
    ]
    for i, (line, desc, val) in enumerate(part1, start=5):
        ws_4562.cell(row=i, column=1, value=line).font = Font(bold=True)
        ws_4562.cell(row=i, column=2, value=desc)
        ws_4562.cell(row=i, column=3, value=val)

    ws_4562["A10"] = "PART II — Special Depreciation Allowance (OBBBA 100% Bonus Depreciation)"
    ws_4562["A10"].font = Font(bold=True, color="4F8EF7")
    ws_4562.cell(row=11, column=1, value="Line 14").font = Font(bold=True)
    ws_4562.cell(row=11, column=2,
        value="100% bonus depreciation — qualifying property placed in service after 2025-01-19")
    ws_4562.cell(row=11, column=3, value=f"${bonus_total:,.2f}")

    ws_4562["A13"] = "PART III — MACRS Depreciation Detail"
    ws_4562["A13"].font = Font(bold=True, color="4F8EF7")
    detail_headers = ["Asset Description", "Date Placed in Service", "Basis ($)",
                      "Recovery Period", "Method", "Deduction Type", "Deductible ($)"]
    for col, h in enumerate(detail_headers, start=1):
        ws_4562.cell(row=14, column=col, value=h)
    _style_header_row(ws_4562, 14, len(detail_headers))

    for ri, d in enumerate(report["deductions"], start=15):
        is_179  = d["deduction_type"] == "Section 179"
        period  = "Immediate" if not is_179 else "Immediate (Sec 179)"
        method  = "Section 179" if is_179 else "100% Bonus Dep (OBBBA)"
        ws_4562.cell(row=ri, column=1, value=d["asset_name"])
        ws_4562.cell(row=ri, column=2, value="See ledger")
        ws_4562.cell(row=ri, column=3, value=float(d["deductible_amount"]))
        ws_4562.cell(row=ri, column=4, value=period)
        ws_4562.cell(row=ri, column=5, value=method)
        ws_4562.cell(row=ri, column=6, value=d["deduction_type"])
        ws_4562.cell(row=ri, column=7, value=float(d["deductible_amount"]))

    total_row = 15 + len(report["deductions"])
    ws_4562.cell(row=total_row, column=1, value="TOTAL DEPRECIATION DEDUCTIONS").font = Font(bold=True)
    ws_4562.cell(row=total_row, column=7,
                 value=sec179_total + bonus_total).font = Font(bold=True)
    _auto_width(ws_4562)

    # ------------------------------------------------------------------ #
    # Sheet 7 — Quarterly Estimated Tax Payments
    # ------------------------------------------------------------------ #
    ws_qtr = wb.create_sheet("Quarterly Estimates")
    ws_qtr["A1"] = f"Quarterly Estimated Tax Payments — {report['tax_year']}"
    ws_qtr["A1"].font = TITLE_FONT
    ws_qtr["A2"] = (
        "IRS Form 1040-ES / 1120-W schedule. "
        "LLC taxed as partnership: pay through partners' individual estimated taxes."
    )
    ws_qtr["A2"].font = MUTED_FONT
    ws_qtr.row_dimensions[2].height = 30

    net_liability = float(report["summary"]["estimated_net_liability"])
    quarterly     = net_liability / 4

    qtr_headers = ["Quarter", "Due Date", "Estimated Payment ($)", "Cumulative ($)", "Status", "Notes"]
    for col, h in enumerate(qtr_headers, start=1):
        ws_qtr.cell(row=4, column=col, value=h)
    _style_header_row(ws_qtr, 4, len(qtr_headers))

    quarters = [
        ("Q1 2026", "April 15, 2026",   quarterly,     quarterly,     "⚠ Due / Past", "1st installment"),
        ("Q2 2026", "June 16, 2026",    quarterly, 2 * quarterly,     "Upcoming",     "2nd installment"),
        ("Q3 2026", "September 15, 2026", quarterly, 3 * quarterly,   "Upcoming",     "3rd installment"),
        ("Q4 2026", "January 15, 2027", quarterly, 4 * quarterly,     "Upcoming",     "4th installment"),
    ]
    for ri, (qtr, due, pmt, cum, status, note) in enumerate(quarters, start=5):
        ws_qtr.cell(row=ri, column=1, value=qtr).font = Font(bold=True)
        ws_qtr.cell(row=ri, column=2, value=due)
        ws_qtr.cell(row=ri, column=3, value=round(pmt, 2))
        ws_qtr.cell(row=ri, column=4, value=round(cum, 2))
        status_cell = ws_qtr.cell(row=ri, column=5, value=status)
        if "Past" in status:
            status_cell.font = Font(color="E05C5C", bold=True)
        ws_qtr.cell(row=ri, column=6, value=note)

    ws_qtr.cell(row=10, column=1, value="TOTAL ANNUAL LIABILITY (Est.)").font = Font(bold=True)
    ws_qtr.cell(row=10, column=3, value=round(net_liability, 2)).font = Font(bold=True)

    ws_qtr["A12"] = "Safe-harbour rule: pay ≥ 100% of prior year tax (110% if AGI > $150k) to avoid underpayment penalties."
    ws_qtr["A12"].font = MUTED_FONT
    ws_qtr["A13"] = "State estimated taxes (NJ, CT, CA) must be filed separately — see Filing Checklist sheet."
    ws_qtr["A13"].font = MUTED_FONT
    _auto_width(ws_qtr)

    # ------------------------------------------------------------------ #
    # Sheet 8 — Filing Checklist (replaces CPA prep session)
    # ------------------------------------------------------------------ #
    ws_chk = wb.create_sheet("Filing Checklist")
    ws_chk["A1"] = f"CPA-Replacement Filing Checklist — Tax Year {report['tax_year']}"
    ws_chk["A1"].font = TITLE_FONT
    ws_chk["A2"] = "Stark Financial Holdings LLC  |  Auto-generated by Stark Tax Engine"
    ws_chk["A2"].font = MUTED_FONT

    chk_headers = ["#", "Form / Action", "Due Date", "Status", "Details"]
    for col, h in enumerate(chk_headers, start=1):
        ws_chk.cell(row=4, column=col, value=h)
    _style_header_row(ws_chk, 4, len(chk_headers))

    checklist = [
        # Federal — entity
        ("1",  "Form 1065 — U.S. Partnership Return",          "Mar 15 (ext. Sep 15)", "Required",  "LLC taxed as partnership; file for all partners."),
        ("2",  "Schedule K-1 — Partner Shares",                "Mar 15",               "Required",  "Issue to each beneficial owner. Includes QBI, gains, depreciation."),
        ("3",  "Form 4562 — Depreciation & Amortization",      "With 1065",            "Generated", f"Sec 179: ${sec179_total:,.0f}  |  Bonus Dep: ${bonus_total:,.0f}."),
        ("4",  "Schedule D — Capital Gains & Losses",          "With 1065",            "Generated", "See Schedule D sheet. Update cost basis before filing."),
        ("5",  "Form 8960 — NIIT (3.8%)",                      "With 1065/1040",       "Required",  "Net investment income > $200k threshold. Compute per partner."),
        ("6",  "Form 8949 — Sales of Capital Assets",          "With 1065",            "Required",  "Detail each sold asset. Basis column flagged N/A — obtain records."),
        ("7",  "Form 8997 — QOZ Investments",                  "If applicable",        "Review",    "Required if gains reinvested into Qualified Opportunity Fund."),
        # Federal — banking / foreign
        ("8",  "FinCEN 114 (FBAR)",                            "Apr 15 (ext. Oct 15)", "Review",    "Required if foreign financial accounts > $10,000 aggregate."),
        ("9",  "Form 8938 (FATCA)",                            "With 1065",            "Review",    "Required if foreign assets > $50k. LME copper warehouse — confirm jurisdiction."),
        # State filings
        ("10", "NJ-1065 — New Jersey Partnership Return",      "Mar 15",               "Required",  "Alpine Estate (Bergen County). NJ source income from property."),
        ("11", "CT-1065 — Connecticut Partnership Return",     "Mar 15",               "Required",  "Farmington Farm (Hartford County). CT Farm Tax exemption — attach documentation."),
        ("12", "CA 565 — California Partnership Return",       "Mar 15",               "Required",  "Beverly Hills Estate (LA County). CA source income from property."),
        # Action items / flags
        ("13", "⚠ Cost Basis Records",                         "Before filing",        "MISSING",   "No cost basis on record for any asset. Obtain purchase confirmations, broker confirms."),
        ("14", "⚠ Q1 2026 Estimated Tax",                      "Apr 15, 2026",         "URGENT",    f"${quarterly:,.0f} federal payment due. Late = 0.5%/month underpayment penalty."),
        ("15", "⚠ Business-Use % for Rolls-Royce Spectre",     "Before filing",        "REQUIRED",  "Mixed personal/business use — apportion depreciation to business-use %."),
        ("16", "Section 199A QBI Election",                    "With 1065",            "Required",  f"20% QBI deduction. 2026 threshold: $203k (single) / $406k (MFJ)."),
        ("17", "CT Farm Conservation Easement",                "Review",               "Optional",  "Farmington Farm: conservation easement deduction may yield significant savings."),
        ("18", "Alpine NJ — Bid Outcome",                      "Pending",              "Monitor",   "Status: pending. If closed, update acquisition date and cost basis."),
        ("19", "Beverly Hills — LTV Credit Facility",          "At close",             "Monitor",   "60% LTV on $100M appraised = $60M facility at SOFR+240 (6.05%). Confirm lender."),
        ("20", "LME Copper — Collateral Agreement",            "On file",              "Review",    "Confirm J.P. Morgan facility terms. Mark-to-market rules may apply."),
    ]

    for ri, row_data in enumerate(checklist, start=5):
        num, form, due, status, details = row_data
        ws_chk.cell(row=ri, column=1, value=num)
        ws_chk.cell(row=ri, column=2, value=form)
        ws_chk.cell(row=ri, column=3, value=due)
        status_cell = ws_chk.cell(row=ri, column=4, value=status)
        ws_chk.cell(row=ri, column=5, value=details)
        if status in ("MISSING", "URGENT"):
            status_cell.font = Font(bold=True, color="E05C5C")
        elif status == "Generated":
            status_cell.font = Font(color="4CAF50")
        elif status == "Required":
            status_cell.font = Font(color="4F8EF7")

    _auto_width(ws_chk)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
